from openpyxl import load_workbook


def compare_tables_to_html(file1, file2, output_file):
    wb1 = load_workbook(filename=file1)
    wb2 = load_workbook(filename=file2)

    sheet1 = wb1.active
    sheet2 = wb2.active

    max_rows = max(sheet1.max_row, sheet2.max_row)
    max_cols = max(sheet1.max_column, sheet2.max_column)

    compared_data = []
    for i in range(max_rows + 1):
        row_data = []
        for j in range(max_cols + 1):
            if i == 0 and j == 0:
                row_data.append('')  # Top left cell is empty
            elif i == 0:
                row_data.append(f'Column {j}')  # Column headers
            elif j == 0:
                row_data.append(f'Row {i}')  # Row headers
            else:
                cell1 = sheet1.cell(row=i, column=j).value if i <= sheet1.max_row and j <= sheet1.max_column else None
                cell2 = sheet2.cell(row=i, column=j).value if i <= sheet2.max_row and j <= sheet2.max_column else None
                if cell1 is None and cell2 is not None:
                    row_data.append(f"<span style='color:green'>{cell2}</span>")  # New cell
                elif cell1 is not None and cell2 is None:
                    row_data.append(f"<span style='color:red'>{cell1}</span>")  # Deleted cell
                elif cell1 != cell2:
                    row_data.append(
                        f"<span style='color:blue'>{cell1}</span> → <span style='color:green'>{cell2}</span>")       # Changed cell
                else:
                    row_data.append(cell1)  # Unchanged cell
        compared_data.append(row_data)

    # 生成HTML表格
    html_content = '<style>table {border-collapse: collapse;} td, th {border: 1px solid black; padding: 5px;}</style>\n'
    html_content += '<table>\n'
    for row in compared_data:
        html_content += '<tr>\n'
        for cell in row:
            html_content += f'<td>{cell}</td>\n'
        html_content += '</tr>\n'
    html_content += '</table>'

    # 将HTML内容写入文件
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)


if __name__ == "__main__":
    compare_tables_to_html('1.xlsm', '2.xlsm', 'diff.html')
